from __future__ import annotations

import csv
import io
import json
from pathlib import Path
from zipfile import BadZipFile

from commit_change_analyzer.models import NormalizedRow, NormalizedTable

KEY_CANDIDATES = ("id", "key", "code", "name")


def _stringify(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _sanitize_headers(raw_headers: list[object], width: int) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for index in range(width):
        base = _stringify(raw_headers[index]) if index < len(raw_headers) else ""
        base = base or f"column_{index + 1}"
        seen[base] = seen.get(base, 0) + 1
        header = base if seen[base] == 1 else f"{base}_{seen[base]}"
        headers.append(header)
    return headers


def _select_key_field(headers: list[str]) -> str | None:
    lowered = {header.lower(): header for header in headers}
    for candidate in KEY_CANDIDATES:
        for lowered_header, header in lowered.items():
            if lowered_header == candidate or lowered_header.endswith(f"_{candidate}") or lowered_header.startswith(f"{candidate}_"):
                return header
    return headers[0] if headers else None


def _row_key(values: dict[str, str], key_field: str | None, row_index: int) -> str:
    if key_field:
        candidate = values.get(key_field, "").strip()
        if candidate:
            return candidate
    return f"row:{row_index}"


def _build_table(source_path: str, table_name: str, raw_rows: list[list[object]]) -> NormalizedTable:
    width = max((len(row) for row in raw_rows), default=0)
    if width == 0:
        return NormalizedTable(source_path=source_path, table_name=table_name, headers=[], rows=[], key_field=None)
    headers = _sanitize_headers(raw_rows[0], width)
    key_field = _select_key_field(headers)
    rows: list[NormalizedRow] = []
    for source_index, raw_row in enumerate(raw_rows[1:], start=2):
        padded_row = [_stringify(raw_row[index]) if index < len(raw_row) else "" for index in range(width)]
        if not any(padded_row):
            continue
        values = dict(zip(headers, padded_row))
        rows.append(
            NormalizedRow(
                row_key=_row_key(values, key_field, source_index),
                values=values,
                source_index=source_index,
            )
        )
    return NormalizedTable(
        source_path=source_path,
        table_name=table_name,
        headers=headers,
        rows=rows,
        key_field=key_field,
    )


def parse_csv_tables(source_path: str, content: bytes) -> list[NormalizedTable]:
    text = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = [list(row) for row in reader]
    return [_build_table(source_path, Path(source_path).name, rows)]


def parse_json_tables(source_path: str, content: bytes) -> list[NormalizedTable]:
    payload = json.loads(content.decode("utf-8-sig", errors="replace"))
    if isinstance(payload, list) and payload and all(isinstance(item, dict) for item in payload):
        keys: list[str] = []
        for item in payload:
            for key in item.keys():
                if key not in keys:
                    keys.append(str(key))
        rows: list[list[object]] = [keys]
        rows.extend([[item.get(key) for key in keys] for item in payload])  # type: ignore[union-attr]
        return [_build_table(source_path, Path(source_path).name, rows)]
    if isinstance(payload, dict):
        rows = [["field", "value"], *[[key, json.dumps(value, ensure_ascii=False, sort_keys=True)] for key, value in payload.items()]]
        return [_build_table(source_path, Path(source_path).name, rows)]
    rows = [["value"], [json.dumps(payload, ensure_ascii=False, sort_keys=True)]]
    return [_build_table(source_path, Path(source_path).name, rows)]


def parse_excel_tables(source_path: str, content: bytes) -> list[NormalizedTable]:
    try:
        from openpyxl import load_workbook
        from openpyxl.utils.exceptions import InvalidFileException
    except ModuleNotFoundError as error:
        raise RuntimeError(
            "Excel analysis requires openpyxl. Install it with 'python -m pip install .[excel]' "
            "or 'python -m pip install openpyxl'."
        ) from error
    try:
        workbook = load_workbook(io.BytesIO(content), data_only=False, read_only=True)
    except (BadZipFile, InvalidFileException, ValueError) as error:
        raise RuntimeError(f"failed to parse Excel content ({error})") from error
    tables: list[NormalizedTable] = []
    for worksheet in workbook.worksheets:
        raw_rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
        tables.append(_build_table(source_path, worksheet.title, raw_rows))
    return tables


def parse_text_tables(source_path: str, content: bytes) -> list[NormalizedTable]:
    text = content.decode("utf-8-sig", errors="replace")
    rows = [["line_number", "content"]]
    rows.extend([[index, line] for index, line in enumerate(text.splitlines(), start=1)])
    return [_build_table(source_path, Path(source_path).name, rows)]


def load_tables(source_path: str, file_type: str, content: bytes) -> tuple[list[NormalizedTable], list[str]]:
    warnings: list[str] = []
    if not content:
        return [], warnings
    if file_type == "csv":
        return parse_csv_tables(source_path, content), warnings
    if file_type == "json":
        try:
            return parse_json_tables(source_path, content), warnings
        except json.JSONDecodeError as error:
            warnings.append(f"{source_path}: invalid JSON content ({error}).")
            return [], warnings
    if file_type == "text":
        return parse_text_tables(source_path, content), warnings
    if file_type == "excel":
        try:
            return parse_excel_tables(source_path, content), warnings
        except RuntimeError as error:
            warnings.append(f"{source_path}: {error}")
            return [], warnings
    if file_type == "excel-legacy":
        warnings.append(f"{source_path}: .xls is not supported in MVP; please convert it to .xlsx for structured diff.")
        return [], warnings
    if file_type == "excel-binary":
        warnings.append(f"{source_path}: .xlsb is not supported in MVP; please convert it to .xlsx for structured diff.")
        return [], warnings
    warnings.append(f"{source_path}: unsupported file type for structured diff.")
    return [], warnings


def write_normalized_tables(tables: list[NormalizedTable], output_dir: Path, label: str) -> list[str]:
    artifact_paths: list[str] = []
    output_dir.mkdir(parents=True, exist_ok=True)
    for table in tables:
        safe_name = "".join(character if character.isalnum() or character in {"-", "_"} else "_" for character in table.table_name)
        target_path = output_dir / f"{label}__{safe_name}.csv"
        with target_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(table.headers)
            for row in table.rows:
                writer.writerow([row.values.get(header, "") for header in table.headers])
        artifact_paths.append(str(target_path))
    return artifact_paths
